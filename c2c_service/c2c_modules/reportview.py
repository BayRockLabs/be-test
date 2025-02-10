from .models import Employee, SowContract, Timesheet, EmployeeEntryTimesheet, EmployeeUnplannedNonbillableHours, Allocation, Estimation
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView
from datetime import datetime, timedelta, timezone
from c2c_modules.utils import has_permission
from django.utils import timezone
import pandas as pd
from django.http import JsonResponse, HttpResponse
from django.db import models
from decimal import Decimal  # Import Decimal
from django.db.models import Q
from c2c_modules.serializer import ReportSowContractSerializer 
import openpyxl
from openpyxl.utils import get_column_letter

class MissingTimesheetView(APIView):
    
    def get_missing_submissions(self, start_date, end_date):
        # Get all employee IDs
        all_employees = Employee.objects.all()
        employee_ids = {emp.employee_source_id: emp.employee_full_name.strip() for emp in all_employees}

        # Calculate week numbers within the given date range
        current_date = start_date
        week_numbers = []

        while current_date <= end_date:
            week_number = current_date.isocalendar()[1]
            year = current_date.isocalendar()[0]
            
            # Store unique week numbers with their corresponding year
            if (year, week_number) not in week_numbers:
                week_numbers.append((year, week_number))
            
            # Move to the next week
            current_date += timedelta(days=7)

        # Constructing the query using Q objects for EmployeeEntryTimesheet
        timesheet_query = models.Q()
        for week in week_numbers:
            timesheet_query |= models.Q(year=week[0], week_number=week[1])

        # Query EmployeeEntryTimesheet for the specified weeks
        submitted_timesheets = EmployeeEntryTimesheet.objects.filter(timesheet_query).values_list('employee_id__employee_source_id', flat=True)

        # Constructing the query using Q objects for EmployeeUnplannedNonbillableHours
        unplanned_query = models.Q()
        for week in week_numbers:
            unplanned_query |= models.Q(year=week[0], week_number=week[1])

        # Query EmployeeUnplannedNonbillableHours for the specified weeks
        unplanned_hours_records = EmployeeUnplannedNonbillableHours.objects.filter(unplanned_query).values_list('employee_id__employee_unplanned_non_billable_entry_id', flat=True)

        # Combine results from both queries
        combined_submitted_ids = set(submitted_timesheets).union(set(unplanned_hours_records))

        # Find employees who did not submit timesheets
        missing_employees = [
            {'employee_id': emp_id, 'employee_name': employee_ids[emp_id]}
            for emp_id in employee_ids.keys() if emp_id not in combined_submitted_ids
        ]

        return missing_employees

    def get(self, request, *args, **kwargs):
        required_roles = ["c2c_super_admin"]
        result = has_permission(request, required_roles)
        if result["status"] != 200:
            return Response({"result": result}, status=status.HTTP_403_FORBIDDEN)
        
        # Define the response type
        response_type = request.GET.get('response_type', 'JSON').lower()

        # Calculate start and end dates for the previous week if no dates are provided
        if 'start_date' in request.GET and 'end_date' in request.GET:
            start_date_str = request.GET.get('start_date')
            end_date_str = request.GET.get('end_date')
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        else:
            today = timezone.now()
            start_date = today - timedelta(days=today.weekday() + 7)  # Start of last week (Monday)
            end_date = start_date + timedelta(days=6)  # End of last week (Sunday)

        # Get missing submissions
        missing_employees = self.get_missing_submissions(start_date, end_date)

        if response_type == 'download':
            return self.generate_excel_response(missing_employees)

        return JsonResponse({'employees': missing_employees}, status=200)

    def generate_excel_response(self, missing_employees):
        """
        Generate an Excel file with a list of missing submissions.
        """
        output = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        output['Content-Disposition'] = f'attachment; filename="missing_submissions.xlsx"'

        df = pd.DataFrame(missing_employees)
        
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Missing Submissions', index=False)

        return output

class EmployeeUtilizationView(APIView):

    def calculate_total_working_hours(self, start_date, end_date):
        if start_date.tzinfo is None:
            start_date = timezone.make_aware(start_date)
        if end_date.tzinfo is None:
            end_date = timezone.make_aware(end_date)
        total_days = (end_date - start_date).days + 1
        working_days = sum(1 for day in range(total_days) 
                           if (start_date + timedelta(days=day)).weekday() < 5)
        return working_days * 8

    def get(self, request, *args, **kwargs):
        required_roles = ["c2c_super_admin"]
        result = has_permission(request, required_roles)
        if result["status"] != 200:
            return Response({"result": result}, status=status.HTTP_403_FORBIDDEN)
        
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        response_type = request.GET.get('response_type', 'JSON').lower()

        if start_date_str is None:
            start_date = datetime(timezone.now().year, 1, 1)
        else:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")

        if end_date_str is None:
            end_date = timezone.now()
        else:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")

        if start_date.tzinfo is None:
            start_date = timezone.make_aware(start_date)
        if end_date.tzinfo is None:
            end_date = timezone.make_aware(end_date)

        contracts = SowContract.objects.filter(
            start_date__lte=end_date,
            end_date__gte=start_date
        )

        utilization_data = {
            '100%': {'count': 0, 'employees': []},
            '75-99%': {'count': 0, 'employees': []},
            '50-74%': {'count': 0, 'employees': []},
            '<50%': {'count': 0, 'employees': []},
            '0%': {'count': 0, 'employees': []}
        }

        employee_hours = {}
        total_working_hours = self.calculate_total_working_hours(start_date, end_date)
        for contract in contracts:
            try:
                timesheets = Timesheet.objects.filter(
                    contract_sow_id=contract,
                    allocation__contract_sow__start_date__lte=end_date.strftime('%Y-%m-%d'), 
                    allocation__contract_sow__end_date__gte=start_date.strftime('%Y-%m-%d')
                )

                if not timesheets.exists():
                    continue

                total_working_hours = self.calculate_total_working_hours(start_date, end_date)

                for timesheet in timesheets:
                    resource_data = timesheet.resource_estimation_data
                    billable_hours = sum(day['hours'] for day in resource_data.get('Estimation_Data', {}).get('daily', [])
                                         if start_date <= timezone.make_aware(datetime.strptime(day['date'], "%d/%m/%Y")) <= end_date)
                    
                    employee_name = timesheet.resource.employee_full_name
                    billability = resource_data.get('billability', 'unknown').lower()
                    #print(employee_name, billable_hours)
                    if employee_name not in employee_hours:
                        employee_hours[employee_name] = {'billable': 0, 'non_billable': 0}

                    if billability == 'billable':
                        employee_hours[employee_name]['billable'] += billable_hours
                    elif billability == 'non-billable':
                        employee_hours[employee_name]['non_billable'] += billable_hours
                    else:
                        employee_hours[employee_name]['unknown'] += billable_hours

            except Exception as e:
                print(e)
                continue

        for employee_name, hours in employee_hours.items():
            if 'billable' in hours : billable_hours  = hours['billable'] 
            else: billable_hours  = 0
            if 'non_billable' in hours : non_billable_hours  = hours['non_billable'] 
            else: non_billable_hours  = 0
            if 'unknown' in hours : unknown_hours  = hours['unknown'] 
            else: unknown_hours  = 0
            utilization_percentage = ((billable_hours + non_billable_hours + unknown_hours) / total_working_hours * 100) if total_working_hours > 0 else 0
            billable_percentage = (billable_hours / total_working_hours * 100) if total_working_hours > 0 else 0
            non_billable_percentage = (non_billable_hours / total_working_hours * 100) if total_working_hours > 0 else 0
            unknown_percentage = (unknown_hours / total_working_hours * 100) if total_working_hours > 0 else 0
            employee_data = {
                'name': employee_name,
                'billable': f"{billable_percentage:.2f}",
                'non-billable': f"{non_billable_percentage:.2f}",
                'unknown': f"{unknown_percentage:.2f}",
                'worked_hours': f"{billable_hours + non_billable_hours+ unknown_hours:.2f}"
            }
            if utilization_percentage == 100:
                utilization_data['100%']['count'] += 1
                utilization_data['100%']['employees'].append(employee_data)
            elif 75 <= utilization_percentage < 100:
                utilization_data['75-99%']['count'] += 1
                utilization_data['75-99%']['employees'].append(employee_data)
            elif 50 <= utilization_percentage < 75:
                utilization_data['50-74%']['count'] += 1
                utilization_data['50-74%']['employees'].append(employee_data)
            elif utilization_percentage < 50 and utilization_percentage > 0:
                utilization_data['<50%']['count'] += 1
                utilization_data['<50%']['employees'].append(employee_data)
            else:
                utilization_data['0%']['count'] += 1
                utilization_data['0%']['employees'].append(employee_data)
        all_employees = {emp.employee_full_name.strip() for emp in Employee.objects.all()}

        utilized_employee_names = set()
        for key in ['100%', '75-99%', '50-74%', '<50%']:
            utilized_employee_names.update([emp['name'] for emp in utilization_data[key]['employees']])

        zero_percent_employees = all_employees - utilized_employee_names
        for employee_name in zero_percent_employees:
            employee_data = {
                'name': employee_name,
                'billable': "0.00",
                'non-billable': "0.00",
                'unknown': "0.00",
                'worked_hours': "0.00"
            }
            utilization_data['0%']['count'] += 1
            utilization_data['0%']['employees'].append(employee_data)

        # for key in utilization_data.keys():
        #     unique_employees = list({emp['name']: emp for emp in utilization_data[key]['employees']}.values())
        #     utilization_data[key]['employees'] = unique_employees

        if response_type == 'download':
            return self.generate_excel_response(utilization_data)

        response_data = [
            {
                'range': key,
                'count': value['count'],
                'employees': value['employees']
            }
            for key, value in utilization_data.items()
        ]
        output_response = {"total_actual_hours": total_working_hours , 'start_date': start_date, 'end_date': end_date, 'data': response_data}
        return JsonResponse(output_response, safe=False, status=status.HTTP_200_OK)


    def generate_excel_response(self, utilization_data):
        """
        Generate an Excel file with separate sheets for each range.
        """
        output = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        output['Content-Disposition'] = f'attachment; filename="utilization_report.xlsx"'

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            for key in utilization_data.keys():
                df = pd.DataFrame({
                    'Employee Names': utilization_data[key]['employee_names'],
                })
                df.to_excel(writer, sheet_name=key[:31], index=False)  # Sheet names are limited to 31 characters
        return output  
    

class FinancialDataView(APIView):

    def get_financial_data(self, start_date, end_date, client_id=None, contract_sow_id=None):

        set_response_flag=''
        # Filter by contract_sow_id if provided
        if contract_sow_id:
            set_response_flag = 'contracts'
            active_contracts = SowContract.objects.filter(Q(uuid=contract_sow_id) | Q(extension_sow_contract=contract_sow_id))
        # Filter by client_id if provided
        elif client_id :
            set_response_flag = 'client'
            active_contracts = SowContract.objects.filter(client_id=client_id)
        else:
            # Query active SOW contracts within the date range
            set_response_flag = 'all'
            active_contracts = SowContract.objects.filter(
                start_date__lte=end_date,
                end_date__gte=start_date
            )

      

        # Initialize totals
        total_revenue = Decimal(0)  # Initialize as Decimal
        total_ctc = Decimal(0)  # Initialize as Decimal

        # Prepare customer and project data structures
        customers_data = {}
        projects_data = []

        for contract in active_contracts:
            total_revenue += contract.total_contract_amount
            
            # Get pricing information to calculate CTC for this contract
            pricing = contract.pricing
            
            if pricing:
                ctc_for_contract = Decimal(pricing.estimated_company_avg_cost) if pricing.estimated_company_avg_cost is not None else Decimal(0)
                total_ctc += ctc_for_contract
            
                # Aggregate data for customers
                customer_id = contract.client.uuid
                if customer_id not in customers_data:
                    customers_data[customer_id] = {
                        'customer_name': contract.client.name,
                        'revenue': Decimal(0),
                        'profit': Decimal(0),
                        'cost_to_company': Decimal(0)
                    }
                customers_data[customer_id]['revenue'] += contract.total_contract_amount
                customers_data[customer_id]['cost_to_company'] += ctc_for_contract  # Add specific CTC for this contract
            
                # Append project data without billable/non-billable resources
                projects_data.append({
                    "project_id": contract.uuid,
                    "project_name": contract.contractsow_name,
                    "revenue": contract.total_contract_amount,
                    "profit": contract.total_contract_amount - ctc_for_contract,  # Calculate profit per project based on its specific CTC
                    "cost_to_company": ctc_for_contract
                })

        # Calculate profit for organization based on total revenue and total CTC
        total_profit = total_revenue - total_ctc

        # Prepare response data structure
        response_data = {}
        if(set_response_flag == 'all'):
            response_data["organization"] = {
                "revenue": total_revenue,
                "profit": total_profit,
                "cost_to_company": total_ctc,
            }
        else:
            response_data["totals"] = {
                "revenue": total_revenue,
                "profit": total_profit,
                "cost_to_company": total_ctc,
            }

        if(set_response_flag == 'all' or set_response_flag == 'client'): 
            response_data["customers"] = [
                {
                    "customer_id": cid,
                    "customer_name": data['customer_name'],
                    "revenue": data['revenue'],
                    "profit": data['revenue'] - data['cost_to_company'],  # Calculate profit per customer based on their specific CTC
                    "cost_to_company": data['cost_to_company']
                }
                for cid, data in customers_data.items()
            ]
        
        if(set_response_flag == 'all' or set_response_flag == 'contracts'):
            response_data["projects"]= projects_data  # Now includes revenue, profit, and cost to company per project without resource counts.
        

        return response_data

    def get(self, request, *args, **kwargs):
        required_roles = ["c2c_super_admin"]
        result = has_permission(request, required_roles)
        if result["status"] != 200:
            return Response({"result": result}, status=status.HTTP_403_FORBIDDEN)
        
        # Define default date range from Jan 1st of current year to today
        today = timezone.now()
        start_date = datetime(today.year, 1, 1)
        end_date = today

        # Get optional parameters from request
        client_id = request.GET.get('client_id')
        contract_sow_id = request.GET.get('contract_sow_id')
        
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str.strip(), "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str.strip(), "%Y-%m-%d")

        # Call method to retrieve financial data
        financial_data = self.get_financial_data(start_date, end_date, client_id, contract_sow_id)

        response_type = request.GET.get('response_type', 'json').lower()

        if response_type == 'download':
            return self.generate_excel_response(financial_data)

        return JsonResponse(financial_data, status=200)

    def generate_excel_response(self, financial_data):
        """
        Generate an Excel file with financial data.
        """
        output = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        output['Content-Disposition'] = f'attachment; filename="financial_data.xlsx"'

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            # Write organization data
            org_df = pd.DataFrame([financial_data["organization"]])
            org_df.to_excel(writer, sheet_name='Organization', index=False)

            # Write customers data
            customers_df = pd.DataFrame(financial_data["customers"])
            customers_df.to_excel(writer, sheet_name='Customers', index=False)

            # Write projects data
            projects_df = pd.DataFrame(financial_data["projects"])
            projects_df.to_excel(writer, sheet_name='Projects', index=False)

        return output



class ResourceCountsView(APIView):

    def get_resource_counts(self, start_date, end_date, client_id=None, sow_contract_id=None):
        
        set_response_flag = ''
        # Filter by sow_contract_id if provided
        
        if sow_contract_id:
            set_response_flag = 'contract'
            active_contracts = active_contracts = SowContract.objects.filter(Q(uuid=sow_contract_id) | Q(extension_sow_contract=sow_contract_id))
        # Filter by client_id if provided
        elif client_id:
            set_response_flag = 'client'
            active_contracts = SowContract.objects.filter(client_id=client_id)
        else:
            set_response_flag = 'all'
        # Query active SOW contracts within the date range
            active_contracts = SowContract.objects.filter(
            start_date__lte=end_date,
            end_date__gte=start_date
        )
    

        # Initialize totals
        total_billable_resources = 0
        total_non_billable_resources = 0

        projects_data = []

        for contract in active_contracts:
            # Count billable resources from Timesheet
            billable_resources = Timesheet.objects.filter(
                allocation__contract_sow=contract,
                resource_estimation_data__billability='Billable'
            ).values('resource').distinct().count()

            # Count non-billable resources from Timesheet
            non_billable_resources = Timesheet.objects.filter(
                allocation__contract_sow=contract,
                resource_estimation_data__billability='Non-Billable'
            ).values('resource').distinct().count()

            # Append project data
            projects_data.append({
                "project_name": contract.contractsow_name,
                "billable_resources": billable_resources,
                "non_billable_resources": non_billable_resources
            })

            # Update total counts
            total_billable_resources += billable_resources
            total_non_billable_resources += non_billable_resources

        # Prepare response data structure
        response_data = {}
        if (set_response_flag == 'all'):
             response_data["organization"]= {
                "total_billable_resources": total_billable_resources,
                "total_non_billable_resources": total_non_billable_resources,
            }
        else:
              response_data["total_resources"]= {
                "total_billable_resources": total_billable_resources,
                "total_non_billable_resources": total_non_billable_resources,
            }
              
        if(set_response_flag == 'all' or set_response_flag == 'client'):
            response_data["projects"]=  projects_data

        if(set_response_flag == 'all' or set_response_flag == 'contract'):
            response_data["projects"]=  projects_data
        
        return response_data

    def get(self, request, *args, **kwargs):
        required_roles = ["c2c_super_admin"]
        result = has_permission(request, required_roles)
        if result["status"] != 200:
            return Response({"result": result}, status=status.HTTP_403_FORBIDDEN)

        # Define default date range from Jan 1st of current year to today
        today = timezone.now()
        start_date = datetime(today.year, 1, 1)
        end_date = today

        # Get optional parameters from request
        client_id = request.GET.get('client_id')
        sow_contract_id = request.GET.get('sow_contract_id')
        
        # Get optional date parameters from request and replace default date range if provided
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str.strip(), "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str.strip(), "%Y-%m-%d")

        # Call method to retrieve resource counts
        resource_counts = self.get_resource_counts(start_date, end_date, client_id, sow_contract_id)

        response_type = request.GET.get('response_type', 'json').lower()

        if response_type == 'download':
            return self.generate_excel_response(resource_counts)

        return JsonResponse(resource_counts, status=200)

    def generate_excel_response(self, resource_counts):
        """
        Generate an Excel file with resource counts.
        """
        output = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        output['Content-Disposition'] = f'attachment; filename="resource_counts.xlsx"'

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            # Write organization data
            org_df = pd.DataFrame([resource_counts["organization"]])
            org_df.to_excel(writer, sheet_name='Organization', index=False)

            # Write projects data
            projects_df = pd.DataFrame(resource_counts["projects"])
            projects_df.to_excel(writer, sheet_name='Projects', index=False)

        return output



class ContractBurndownView(APIView):
    
    def get_current_estimation_data(self, estimation_list):
        new_estimation_indexes = []
        for resource in estimation_list:
            for resource_count in range(0, resource['num_of_resources']):
                new_estimation_indexes.append(resource)
        return new_estimation_indexes
    
    def get(self, request, contract_id):
        # Step 1: Retrieve the SOW contract
        try:
            sow_contract = SowContract.objects.get(uuid=contract_id)
        except SowContract.DoesNotExist:
            return Response({"error": "Contract not found."}, status=status.HTTP_404_NOT_FOUND)

        # Step 2: Get extensions if applicable
        extension_contracts = []

        if sow_contract.extension_sow_contract not in [None, "", "null"]:
            # Only query if extension_sow_contract is not null
            try:
                extension_contracts = SowContract.objects.filter(uuid=sow_contract.extension_sow_contract)
            except ValueError:
                print({"error": "Invalid extension contract ID."})
        else:
            print("No extension contract to process")

        # Step 3: Get all allocations for this contract and its extensions
        allocations = Allocation.objects.filter(contract_sow__in=[sow_contract] + list(extension_contracts))

        # Prepare data structure for response
        result = []
        
        # Get start and end dates from the primary contract
        start_date = datetime.strptime(sow_contract.start_date, "%Y-%m-%d")  # Adjust format as necessary
        end_date = datetime.strptime(sow_contract.end_date, "%Y-%m-%d") if sow_contract.end_date else timezone.now()

        current_date = start_date
        
        # Step 4: Calculate weekly budgets and actual costs
        while current_date <= end_date:
            week_number = current_date.isocalendar()[1]
            planned_budget = 0
            actual_cost = 0
            allocation_cost = 0

            # Calculate planned budget for this week
            for allocation in allocations:
                for resource in allocation.resource_data:
                    resource_id = resource.get("resource_id")
                    if not resource_id:
                        continue
                    
                    # Check if resource_id is "Budgeto123"
                    if resource_id == "BUDGETO123":
                        estimation_id = allocation.estimation_id
                        try:
                            estimation_record = Estimation.objects.get(uuid=estimation_id)  # Fetching estimation record
                            estimation_resource_data = estimation_record.resource
                            # Call function to get current estimation data
                            new_estimation_indexes = self.get_current_estimation_data(estimation_resource_data)
                            
                            # Compare with existing resources and pick matching index (implement logic as needed)
                            for index, resource in enumerate(allocation.resource_data):
                                # Check if resource_id matches 'BUDGETO123'
                                if resource.get('resource_id') == 'BUDGETO123':
                                    # Fetch the corresponding item from new_estimation_indexes
                                    estimation_data = new_estimation_indexes[index]
                                    # Get daily_entries from estimation_data
                                    daily_entries = estimation_data.get("Estimation_Data", {}).get("daily", [])
                                    # Calculate total_hours_for_week
                                    total_hours_for_week = sum(
                                        entry['hours'] for entry in daily_entries 
                                        if datetime.strptime(entry['date'], '%d/%m/%Y').date().isocalendar()[1] == week_number
                                    )
                                    # Get the bill_rate from estimation_data or set default to 0
                                    bill_rate = estimation_data.get("pay_rate_info", {}).get("billrate", 0)
                                    bill_rate = float(bill_rate)  
                                    # Calculate the planned_budget
                                    allocation_cost += total_hours_for_week * bill_rate

                        except Estimation.DoesNotExist:
                            print(f"No estimation found for ID {estimation_id}")

                    # Query Timesheet for this resource_id
                    timesheets = Timesheet.objects.filter(resource_id=resource_id)
                    
                    # Calculate total hours for this week from Estimation_Data in Timesheet
                    for timesheet in timesheets:
                        estimation_data = timesheet.resource_estimation_data.get("Estimation_Data", {})
                        
                        daily_entries = estimation_data.get("daily", [])
                        
                        total_hours_for_week = sum(
                            entry['hours'] for entry in daily_entries 
                            if datetime.strptime(entry['date'], '%d/%m/%Y').date().isocalendar()[1] == week_number
                        )
                        
                        # bill_rate = estimation_data.get("pay_rate_info", {}).get("billrate", 0)  # Get bill rate
                        bill_rate = timesheet.resource_estimation_data.get("pay_rate_info", {}).get("billrate", 0)
                        
                        planned_budget += total_hours_for_week * bill_rate
                        
                        entry_timesheet_records = EmployeeEntryTimesheet.objects.filter(
                            timesheet_id=timesheet,
                            week_number=week_number,
                            employee_id=resource_id  # Ensure we match with employee ID
                        )
                        
                        for entry in entry_timesheet_records:
                            if entry.ts_approval_status == "submitted":  # Only consider submitted timesheets
                                actual_cost += entry.billable_hours * timesheet.resource_estimation_data.get("pay_rate_info", {}).get("billrate", 0)
                        
            remaining_budget = planned_budget - actual_cost
            
            result.append({
                "week": week_number,
                "planned_budget": planned_budget,
                "actual_cost": actual_cost,
                "allocation_cost": allocation_cost,
                "remaining_budget": remaining_budget,
                "week_start_date": current_date,
                "week_end_date": current_date + timedelta(days=4)  # Assuming a work week of Monday to Friday
            })
            
            # Move to next week
            current_date += timedelta(weeks=1)

        return Response(result, status=status.HTTP_200_OK)

from datetime import datetime, timedelta
import calendar
from collections import defaultdict
class SowContractAPIView(APIView):
    def get(self, request, contract_id):
        contracts = SowContract.objects.filter(Q(extension_sow_contract=contract_id) | Q(uuid=contract_id))
        if not contracts.exists():
            return Response({"error": "No contracts found"}, status=status.HTTP_404_NOT_FOUND)
        total_project_amount = sum(contract.total_contract_amount for contract in contracts)
        allocations = Allocation.objects.filter(contract_sow__in=contracts)
        timesheets = Timesheet.objects.filter(allocation__in=allocations)
        timesheet_data = []
        for timesheet in timesheets:
            employee_weekly_data = self.get_timesheet_data(timesheet)
            timesheet_data.extend(employee_weekly_data)
        data = self.aggregate_weekly_data(timesheet_data)
        contract_week_data = self.get_contract_weekly_data(contracts)
        output_data = self.compare_week_data(contract_week_data, data)
        remaining_amount = total_project_amount
        graph_info_list = []
        for item in output_data:
            graph_info_dict = dict()
            remaining_amount = float(remaining_amount)
            graph_info_dict["allocated_budget"] = item["allocated_amount"]
            graph_info_dict["actual_spent"] = item["billable_amount"] + item["non_billable_amount"]
            graph_info_dict["remaining_allocated_budget"] = remaining_amount - graph_info_dict["allocated_budget"]
            graph_info_dict["leftover_budget"] = remaining_amount - graph_info_dict["actual_spent"]
            graph_info_dict["week_start_date"] = item["week_start_date"]
            graph_info_dict["week_end_date"] = item["week_end_date"]
            graph_info_dict["week_number"] = item["week_number"]
            graph_info_dict["year"] = item["year"]
            remaining_amount = graph_info_dict["remaining_allocated_budget"]
            graph_info_list.append(graph_info_dict)
        
        response_data = {
            "total_project_amount": total_project_amount,
            "comparison_data": graph_info_list,
        }
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    
    def get_contract_weekly_data(self, contracts):
        data = []
        unique_weeks = set()
        for contract in contracts:
            contract_start_date = contract.start_date
            contract_end_date = contract.end_date
            weeks = self.get_weeks_between_dates(contract_start_date, contract_end_date)
            for week in weeks:
                week_key = (week["week_number"], week["year"])
                if week_key not in unique_weeks:
                    unique_weeks.add(week_key)
                    data.append(week)
        data.sort(key=lambda x: (x["year"], x["week_number"]))
        return data
    
    def compare_week_data(self, contract_week_data, timesheets_count):
        timesheet_dict = {}
        for entry in timesheets_count:
            timesheet_key = (entry["week_number"], entry["year"])
            timesheet_dict[timesheet_key] = {
                "allocated_amount": entry["allocated_amount"],
                "allocated_hours": entry["allocated_hours"],
                "billable_amount": entry["billable_amount"],
                "billable_hours": entry["billable_hours"],
                "non_billable_amount": entry["non_billable_amount"],
                "non_billable_hours": entry["non_billable_hours"]
            }
        result = []
        for contract in contract_week_data:
            contract_key = (contract["week_number"], contract["year"])
            if contract_key in timesheet_dict:
                contract_data = {
                    "week_number": contract["week_number"],
                    "year": contract["year"],
                    "week_start_date": contract["week_start_date"],
                    "week_end_date": contract["week_end_date"],
                    "allocated_amount": timesheet_dict[contract_key]["allocated_amount"],
                    "allocated_hours": timesheet_dict[contract_key]["allocated_hours"],
                    "billable_amount": timesheet_dict[contract_key]["billable_amount"],
                    "billable_hours": timesheet_dict[contract_key]["billable_hours"],
                    "non_billable_amount": timesheet_dict[contract_key]["non_billable_amount"],
                    "non_billable_hours": timesheet_dict[contract_key]["non_billable_hours"]
                }
            else:
                contract_data = {
                    "week_number": contract["week_number"],
                    "year": contract["year"],
                    "week_start_date": contract["week_start_date"],
                    "week_end_date": contract["week_end_date"],
                    "allocated_amount": 0,
                    "allocated_hours": 0,
                    "billable_amount": 0,
                    "billable_hours": 0,
                    "non_billable_amount": 0,
                    "non_billable_hours": 0
                }
            result.append(contract_data)
        return result

    def get_weeks_between_dates(self, start_date_str, end_date_str):
        weeks = []
        current_date = datetime.today().date()
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        start_date = start_date.date() if isinstance(start_date, datetime) else start_date
        end_date = end_date.date() if isinstance(end_date, datetime) else end_date
        while start_date <= end_date:
            week_number = start_date.isocalendar()[1]
            year = start_date.isocalendar()[0]
            if start_date <= current_date <= end_date:
                reference_week = current_date.isocalendar()[1]
                reference_year = current_date.isocalendar()[0]
            else:
                reference_week = week_number
                reference_year = year
            week_start_date = start_date - timedelta(days=start_date.weekday())
            week_end_date = week_start_date + timedelta(days=4)
            if week_end_date > end_date:
                week_end_date = end_date
            week_data = {
                "week_number": reference_week,
                "year": reference_year,
                "week_start_date": week_start_date.strftime("%Y-%m-%d"),
                "week_end_date": week_end_date.strftime("%Y-%m-%d"),
            }
            if week_data not in weeks:
                weeks.append(week_data)
            start_date = week_start_date + timedelta(days=7)
        weeks.sort(key=lambda x: (x["year"], x["week_number"]))
        return weeks
    def aggregate_weekly_data(self,timesheet_data):
        result = []
        aggregated_data = defaultdict(lambda: {
                            "week_start_date": None,
                            "week_end_date": None,
                            "week_number": None,
                            "year": None,
                            "allocated_hours": 0,
                            "allocated_amount": 0,
                            "billable_amount": 0,
                            "billable_hours": 0,
                            "non_billable_amount": 0,
                            "non_billable_hours": 0
                        })
        for entry in timesheet_data:
            key = (entry["week_number"], entry["year"])
            
            if aggregated_data[key]["week_start_date"] is None:
                aggregated_data[key]["week_start_date"] = entry["week_start_date"]
                aggregated_data[key]["week_end_date"] = entry["week_end_date"]
                aggregated_data[key]["week_number"] = entry["week_number"]
                aggregated_data[key]["year"] = entry["year"]
            aggregated_data[key]["allocated_amount"] += entry["allocated_amount"]
            aggregated_data[key]["allocated_hours"] += entry["allocated_hours"]
            aggregated_data[key]["billable_amount"] += entry["billable_amount"]
            aggregated_data[key]["billable_hours"] += entry["billable_hours"]
            aggregated_data[key]["non_billable_amount"] += entry["non_billable_amount"]
            aggregated_data[key]["non_billable_hours"] += entry["non_billable_hours"]
        result = [
        {
            "week_number": week,
            "year": year,
            "week_start_date": values["week_start_date"],
            "week_end_date": values["week_end_date"],
            "allocated_amount": values["allocated_amount"],
            "allocated_hours": values["allocated_hours"],
            "billable_amount": values["billable_amount"],
            "billable_hours": values["billable_hours"],
            "non_billable_amount": values["non_billable_amount"],
            "non_billable_hours": values["non_billable_hours"]
        }
        for (week, year), values in sorted(aggregated_data.items())
            ]
        return result

    def get_timesheet_data(self, timesheet):
        resource_estimation_data = timesheet.resource_estimation_data
        timesheet_id = timesheet.id
        employee_id = timesheet.resource_id
        daily_entries = resource_estimation_data.get("Estimation_Data",[]).get("weekly", [])
        weekly_data = self.process_weekly_data(daily_entries,timesheet_id,employee_id)
        weekly_data = self.get_billable_hours(resource_estimation_data, weekly_data)
        return weekly_data
    
    def get_billable_hours(self, resource_estimation_data,weekly_data):
        for data in weekly_data:
            record = EmployeeEntryTimesheet.objects.filter(timesheet_id=data["timesheet_id"], week_number=data["week_number"], employee_id=data["employee_id"], year=data["year"],ts_approval_status="approved")
            allocated_amount = data["allocated_hours"] * resource_estimation_data.get("pay_rate_info", {}).get("billrate", 0)
            if record.exists():
                billable_hours = record.billable_hours
                non_billable_hours = record.non_billable_hours
                billable_amount = billable_hours * resource_estimation_data.get("pay_rate_info", {}).get("billrate", 0)
                non_billable_amount = non_billable_hours * resource_estimation_data.get("pay_rate_info", {}).get("billrate", 0)
                data["allocated_amount"] = allocated_amount
                data["billable_amount"] = billable_amount
                data["non_billable_amount"] = non_billable_amount
                data["billable_hours"] = billable_hours
                data["non_billable_hours"] = non_billable_hours
            else:
                data["allocated_amount"] = allocated_amount
                data["billable_amount"] = 0
                data["non_billable_amount"] = 0
                data["billable_hours"] = 0
                data["non_billable_hours"] = 0
        return weekly_data
    def get_week_start_end(self, week_number, year):
        first_day_of_year = datetime(year, 1, 1)
        days_to_add = timedelta(weeks=week_number-1)
        start_date = first_day_of_year + days_to_add - timedelta(days=first_day_of_year.weekday())
        end_date = start_date + timedelta(days=4)
        return start_date, end_date

    def process_weekly_data(self, weekly_data, timesheet_id, employee_id):
        week_data = []
        for entry in weekly_data:
            week_number = entry["week"]
            year = 2025
            week_hours = entry["hours"]
            week_start_date, week_end_date = self.get_week_start_end(week_number, year)
            week_data.append({
                "week_number": week_number,
                "week_start_date": week_start_date.strftime("%d/%m/%Y"),
                "week_end_date": week_end_date.strftime("%d/%m/%Y"),
                "allocated_hours": week_hours,
                "year": year,
                "timesheet_id": timesheet_id,
                "employee_id": employee_id,
            })
        return week_data


class ContractsEndingReportAPIView(APIView):
    """
    API view to retrieve contracts ending within a given number of weeks along with related timesheets.
    """
    def post(self, request, *args, **kwargs):
        weeks = request.data.get("weeks", 2)
        export_type = request.data.get("export_type", "json")
        today = timezone.now().date()
        contracts = SowContract.objects.exclude(end_date__isnull=True).exclude(end_date="")
        closing_data = {f"Closed_in_week_{week}": [] for week in range(1, weeks + 1)}
        for contract in contracts:
            try:
                end_date = datetime.strptime(contract.end_date, "%Y-%m-%d").date()
            except ValueError:
                continue
            
            for week in range(1, weeks + 1):
                if today + timedelta(days=(week - 1) * 7) < end_date <= today + timedelta(days=week * 7):
                    related_timesheets = Timesheet.objects.filter(contract_sow=contract).select_related("resource", "client")
                    
                    for timesheet in related_timesheets:
                        closing_data[f"Closed_in_week_{week}"].append({
                            "employee_id": timesheet.resource.employee_source_id,
                            "employee_name": timesheet.resource.employee_full_name,
                            "contract_name": contract.contractsow_name,
                            "contract_start_date": contract.start_date,
                            "contract_end_date": contract.end_date,
                            "client_name": timesheet.client.name if timesheet.client else None,
                            "status": f"Completing contract in {week} week(s)"
                        })
                    break  # Stop checking further weeks once assigned
        if export_type == "excel":
            return self.export_contracts_to_excel(closing_data,weeks)
        else:
            return Response(
                {
                    "placement_ending_report": closing_data,
                    "closing_count_weeks": weeks
                },
                status=status.HTTP_200_OK
            )
    
    def export_contracts_to_excel(self, placement_ending_report, closing_count_weeks):
        """
        Exports the placement ending report data to an Excel file.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove the default sheet

        # Define column headers
        headers = ["Employee ID", "Employee Name", "Contract Name", "Start Date", "End Date", "Client Name", "Status"]

        # Create a sheet for each closing week
        for week_num in range(1, closing_count_weeks + 1):
            week_key = f"Closed_in_week_{week_num}"
            if week_key not in placement_ending_report:
                continue
            
            ws = wb.create_sheet(title=f"Week {week_num}")
            
            # Write headers
            ws.append(headers)

            # Write contract data
            for contract in placement_ending_report[week_key]:
                ws.append([
                    contract["employee_id"],
                    contract["employee_name"],
                    contract["contract_name"],
                    contract["contract_start_date"],
                    contract["contract_end_date"],
                    contract["client_name"],
                    contract["status"],
                ])

            # Auto-adjust column widths
            for col_idx, _ in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].auto_size = True

        # Prepare response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="contracts_report.xlsx"'

        wb.save(response)
        return response